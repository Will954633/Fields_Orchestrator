#!/usr/bin/env python3
"""
Sync the mini-site health audit to a Google Sheet (in a target Drive folder).

  Tab 1 "Dashboard"  — one row per home (health %, status counts, freshness, slug).
  Tabs 2..N          — one per home, every dynamic field with value / status /
                       last updated / last changed / errors.

Mechanism: build a multi-tab .xlsx with openpyxl (status cells colour-coded), then
import it into Drive as a Google Sheet. Reuses one file ("Mini-Site Health") in the
folder so the URL is stable across daily syncs. Uses Drive API only (the Sheets API
is disabled on the OAuth project; the full auth/drive scope covers Drive import).

Usage:
  python3 scripts/minisite_health_to_sheet.py            # build/refresh
  python3 scripts/minisite_health_to_sheet.py --no-snapshot
"""
from __future__ import annotations
import argparse
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import minisite_health_check as hc

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

FOLDER_ID = "1x6GGEubGPhsPMUhaN8mKUpLBTHEVZs11"
SHEET_NAME = "Mini-Site Health"
TOKEN_FILE = "/home/fields/.gdrive-server-credentials.json"
KEYS_FILE = "/home/fields/.gdrive-oauth.keys.json"
XLSX_PATH = "/tmp/minisite_health.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

STATUS_FILL = {
    "OK":                "D9EFD4",  # green
    "STALE":             "FFE699",  # amber
    "MISSING":           "F5CCCC",  # red
    "ERROR":             "EB9999",  # strong red
    "PENDING-EXPECTED":  "E6E6E6",  # grey
    "UNKNOWN-FRESHNESS": "D9D9F2",  # blue-grey
    "KNOWN-GAP":         "EDEDED",  # light grey
}
HEADER_FILL = PatternFill("solid", fgColor="33414D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=12)

DASH_HEADERS = ["Address", "Suburb", "State", "Entered", "Built", "Gap", "Health %",
                "Errors", "Missing", "Stale", "Pending", "Unknown", "Last data pull", "Slug"]
HOME_HEADERS = ["Tab", "Field", "Value", "Status", "Freshness source",
                "Last updated", "Last changed", "Detail / error", "Note"]


# ---- auth ---------------------------------------------------------------------
def get_drive():
    tok = json.load(open(TOKEN_FILE))
    keys = json.load(open(KEYS_FILE))
    k = keys.get("installed") or keys.get("web")
    creds = Credentials(
        token=tok.get("access_token"),
        refresh_token=tok.get("refresh_token"),
        token_uri=k.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=k["client_id"], client_secret=k["client_secret"],
        scopes=[tok.get("scope", "https://www.googleapis.com/auth/drive")],
    )
    if not creds.valid:
        creds.refresh(Request())
        tok["access_token"] = creds.token
        if creds.expiry:
            tok["expiry_date"] = int(creds.expiry.timestamp() * 1000)
        json.dump(tok, open(TOKEN_FILE, "w"))
    return build("drive", "v3", credentials=creds)


# ---- helpers ------------------------------------------------------------------
def fmt_ts(iso):
    return (iso or "")[:19].replace("T", " ")


def excel_title(slug, used):
    """Excel tab titles: <=31 chars, no []:*?/\\, unique."""
    t = slug[:31]
    if t in used:
        base = slug[:28]
        i = 1
        while f"{base}_{i}" in used:
            i += 1
        t = f"{base}_{i}"
    used.add(t)
    return t


def style_header_block(ws, header_row, ncols):
    ws.freeze_panes = f"A{header_row + 1}"
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=1, column=1).font = TITLE_FONT


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---- workbook -----------------------------------------------------------------
def build_workbook(results, now_utc):
    wb = Workbook()
    used = set()

    # Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    ws["A1"] = (f"Mini-Site Health — generated "
                f"{now_utc.astimezone(hc.AEST):%Y-%m-%d %H:%M AEST}  ·  "
                f"expected last nightly run "
                f"{hc.expected_last_run(now_utc).astimezone(hc.AEST):%Y-%m-%d %H:%M}")
    for c, h in enumerate(DASH_HEADERS, start=1):
        ws.cell(row=2, column=c, value=h)
    for r in results:
        cc = r["counts"]
        ws.append([
            r["address"], r["suburb"], r["state"] or "",
            r.get("entered") or "—", r.get("built") or "—", r.get("gap") or "—",
            r["health_pct"],
            cc.get("ERROR", 0), cc.get("MISSING", 0), cc.get("STALE", 0),
            cc.get("PENDING-EXPECTED", 0), cc.get("UNKNOWN-FRESHNESS", 0),
            fmt_ts(str(r["data_pull_date"])), r["slug"],
        ])
    # colour Gap cell (col 6) by gap_status, Health % cell (col 7) by threshold
    for i, r in enumerate(results, start=3):
        gfill = STATUS_FILL.get(r.get("gap_status"))
        if gfill:
            ws.cell(row=i, column=6).fill = PatternFill("solid", fgColor=gfill)
        hp = r["health_pct"]
        colour = "D9EFD4" if hp >= 90 else "FFE699" if hp >= 70 else "F5CCCC"
        ws.cell(row=i, column=7).fill = PatternFill("solid", fgColor=colour)
    style_header_block(ws, 2, len(DASH_HEADERS))
    autofit(ws, [42, 16, 13, 16, 16, 9, 9, 7, 8, 7, 9, 9, 19, 34])

    # Per-home tabs
    title_map = {}
    for r in results:
        t = excel_title(r["slug"], used)
        title_map[r["slug"]] = t
        ws = wb.create_sheet(title=t)
        ws["A1"] = (f"{r['address']}  ·  {r['suburb']}  ·  state={r['state']}  ·  "
                    f"health {r['health_pct']}%")
        for c, h in enumerate(HOME_HEADERS, start=1):
            ws.cell(row=2, column=c, value=h)
        status_col = HOME_HEADERS.index("Status") + 1
        row = 3
        for fr in r["fields"]:
            ws.append([
                fr["tab"], fr["field"], str(fr["value"]), fr["status"],
                fr["freshness_src"] or "", fmt_ts(fr["freshness_ts"]),
                fmt_ts(fr["last_changed"]), fr["detail"] or "", fr["note"] or "",
            ])
            fill = STATUS_FILL.get(fr["status"])
            if fill:
                ws.cell(row=row, column=status_col).fill = PatternFill("solid", fgColor=fill)
            row += 1
        style_header_block(ws, 2, len(HOME_HEADERS))
        autofit(ws, [12, 26, 40, 18, 20, 19, 19, 44, 30])

    wb.save(XLSX_PATH)
    return title_map


# ---- drive upload -------------------------------------------------------------
def find_existing(drive):
    q = (f"name='{SHEET_NAME}' and '{FOLDER_ID}' in parents "
         f"and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false")
    hits = drive.files().list(q=q, fields="files(id,name)", supportsAllDrives=True,
                              includeItemsFromAllDrives=True).execute().get("files", [])
    return hits[0]["id"] if hits else None


def upload(drive):
    media = MediaFileUpload(XLSX_PATH, mimetype=XLSX_MIME, resumable=False)
    ssid = find_existing(drive)
    if ssid:
        drive.files().update(fileId=ssid, media_body=media, supportsAllDrives=True).execute()
    else:
        body = {"name": SHEET_NAME, "parents": [FOLDER_ID],
                "mimeType": "application/vnd.google-apps.spreadsheet"}
        ssid = drive.files().create(body=body, media_body=media, fields="id",
                                    supportsAllDrives=True).execute()["id"]
    return ssid


# ---- main ---------------------------------------------------------------------
def set_env_from_file():
    if os.environ.get("COSMOS_CONNECTION_STRING"):
        return
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env")
    if os.path.exists(env_path):
        for line in open(env_path):
            if "=" in line and not line.startswith("#"):
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip().strip('"'))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-snapshot", action="store_true")
    args = ap.parse_args()

    set_env_from_file()
    results, now_utc = hc.run_audit(persist=not args.no_snapshot)
    build_workbook(results, now_utc)
    drive = get_drive()
    ssid = upload(drive)
    print(f"Synced {len(results)} reports → "
          f"https://docs.google.com/spreadsheets/d/{ssid}/edit")


if __name__ == "__main__":
    main()
