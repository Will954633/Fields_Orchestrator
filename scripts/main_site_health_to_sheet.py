#!/usr/bin/env python3
"""
Sync the business-wide "Fields Systems Health" audit to a Google Sheet (in a
target Drive folder) — the single source of truth for every process the
business runs: website data freshness/correctness, the orchestrator pipeline,
the cron/systemd fleet, off-VM GitHub Actions, leads/CRM syncs, ad-decision
logging compliance, and the mini-site (merged in 2026-07-22, was a separate
sheet — see minisite_health_to_sheet.py).

  Tab "Dashboard"          — one row per main-site page (health %, status counts, worst status).
  Tabs (main-site pages)   — one per page, every dynamic data point with value / status /
                             collection / freshness field / last updated / last changed / detail.
  Tab "Mini-Site Dashboard" + one tab per home — same shape, appended from
                             minisite_health_check.py via minisite_health_to_sheet.build_workbook(wb=...).

Mechanism: build a multi-tab .xlsx with openpyxl (status cells colour-coded), then import
it into Drive as a Google Sheet. Reuses one file (id KNOWN_SHEET_ID below) in the folder so
the URL is stable across daily syncs regardless of the display name. Uses Drive API only
(the Sheets API is disabled on the OAuth project; the full auth/drive scope covers Drive import).

Telegram one-liner is pushed only when something is in breach (ERROR/STALE/MISSING > 0)
across EITHER the main-site or mini-site rows; silent on healthy runs. The Sheet itself
is the primary daily signal.

Usage:
  python3 scripts/main_site_health_to_sheet.py            # build/refresh + alert on breach
  python3 scripts/main_site_health_to_sheet.py --no-snapshot
  python3 scripts/main_site_health_to_sheet.py --no-alert
  python3 scripts/main_site_health_to_sheet.py --no-minisite  # debugging only
"""
from __future__ import annotations
import argparse
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import main_site_health_check as hc

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

FOLDER_ID = "1x6GGEubGPhsPMUhaN8mKUpLBTHEVZs11"
# Renamed 2026-07-22 from "Main Site Health" — this sheet grew from a
# website-only freshness audit into the single business-wide status board
# (Process Registry, GitHub Actions, Market Signals Fetch, Leads & CRM, Ads &
# Compliance, and the mini-site tabs merged in below). Same file id/URL.
SHEET_NAME = "Fields Systems Health"
TOKEN_FILE = "/home/fields/.gdrive-server-credentials.json"
KEYS_FILE = "/home/fields/.gdrive-oauth.keys.json"
XLSX_PATH = "/tmp/main_site_health.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

STATUS_FILL = {
    "OK":                "D9EFD4",  # green
    "STALE":             "FFE699",  # amber
    "MISSING":           "F5CCCC",  # red
    "ERROR":             "EB9999",  # strong red
    "UNKNOWN-FRESHNESS": "D9D9F2",  # blue-grey
    "KNOWN-GAP":         "EDEDED",  # light grey
}
HEADER_FILL = PatternFill("solid", fgColor="33414D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=12)

DASH_HEADERS = ["Page", "Health %", "Error", "Stale", "Missing", "Unknown", "Known-gap",
                "Oldest freshness", "Worst status"]
PAGE_HEADERS = ["Data point", "Scope", "Value", "Status", "Freshness field",
                "Last updated", "Last changed", "Detail"]
SEV_LABEL = {4: "ERROR", 3: "MISSING", 2: "STALE", 1: "UNKNOWN-FRESHNESS", 0: "OK"}


# ---- auth ----------------------------------------------------------------------
# Service account first (same SA api_health_monitor.py uses; the sheet is shared
# with it as editor). The user OAuth refresh token was revoked 2026-06 and broke
# the nightly sync silently for 4 days — SA keys don't expire that way. The
# legacy OAuth path is kept as a fallback only.
SA_KEY_DEFAULT = "/home/fields/.gcp-floor-plan-vision.json"
KNOWN_SHEET_ID = "1Oa7uZv0shzsxftDYJJ3WErxhr7OZMf_SOxRFawbSgTk"


def get_drive():
    sa_key = os.environ.get("GOOGLE_VISION_SA_KEY", SA_KEY_DEFAULT)
    if os.path.exists(sa_key):
        try:
            from google.oauth2 import service_account
            creds = service_account.Credentials.from_service_account_file(
                sa_key, scopes=["https://www.googleapis.com/auth/drive"])
            return build("drive", "v3", credentials=creds)
        except Exception as e:
            print(f"(service-account auth failed: {e} — falling back to OAuth token)")
    return _get_drive_oauth()


def _get_drive_oauth():
    tok = json.load(open(TOKEN_FILE))
    keys = json.load(open(KEYS_FILE))
    k = keys.get("installed") or keys.get("web")
    creds = Credentials(
        token=tok.get("access_token"),
        refresh_token=tok.get("refresh_token"),
        token_uri=k.get("token_uri", "https://oauth2.googleapis.com/token"),
        client_id=k["client_id"], client_secret=k["client_secret"],
        scopes=[tok.get("scope", "https://www.googleapis.com/auth/drive")],
    )
    if not creds.valid:
        creds.refresh(Request())
        tok["access_token"] = creds.token
        if creds.expiry:
            tok["expiry_date"] = int(creds.expiry.timestamp() * 1000)
        json.dump(tok, open(TOKEN_FILE, "w"))
    return build("drive", "v3", credentials=creds)


# ---- helpers ------------------------------------------------------------------
def fmt_ts(iso):
    return (iso or "")[:19].replace("T", " ") if iso else "—"


def style_header_block(ws, header_row, ncols):
    ws.freeze_panes = f"A{header_row + 1}"
    for c in range(1, ncols + 1):
        cell = ws.cell(row=header_row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.cell(row=1, column=1).font = TITLE_FONT


def autofit(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def excel_title(name, used):
    safe = name
    for ch in '[]:*?/\\':
        safe = safe.replace(ch, "-")
    t = safe[:31]
    if t in used:
        i = 1
        while f"{safe[:28]}_{i}" in used:
            i += 1
        t = f"{safe[:28]}_{i}"
    used.add(t)
    return t


# ---- workbook -----------------------------------------------------------------
def build_workbook(pages, now_utc, totals):
    """Builds the main-site tabs into a new Workbook and returns it (unsaved) —
    main() appends the mini-site tabs onto this same `wb` before saving once,
    so both sheets live in one file/one upload/one Telegram digest."""
    wb = Workbook()
    used = set()
    C = hc

    # Dashboard
    ws = wb.active
    ws.title = "Dashboard"
    overall = round(100 * totals.get("OK", 0) /
                    max(1, sum(v for k, v in totals.items() if k != "KNOWN-GAP")))
    ws["A1"] = (f"Fields Systems Health — generated "
                f"{now_utc.astimezone(C.AEST):%Y-%m-%d %H:%M AEST}  ·  overall {overall}%  ·  "
                f"expected last nightly run "
                f"{C.expected_last_run(now_utc).astimezone(C.AEST):%Y-%m-%d %H:%M}")
    for c, h in enumerate(DASH_HEADERS, start=1):
        ws.cell(row=2, column=c, value=h)
    for p in pages:
        cc = p["counts"]
        ws.append([
            p["page"], p["health_pct"],
            cc.get("ERROR", 0), cc.get("STALE", 0), cc.get("MISSING", 0),
            cc.get("UNKNOWN-FRESHNESS", 0), cc.get("KNOWN-GAP", 0),
            fmt_ts(p["oldest_fresh"]), SEV_LABEL.get(p["worst_severity"], "OK"),
        ])
    for i, p in enumerate(pages, start=3):
        hp = p["health_pct"]
        colour = "D9EFD4" if hp >= 90 else "FFE699" if hp >= 70 else "F5CCCC"
        ws.cell(row=i, column=2).fill = PatternFill("solid", fgColor=colour)
        wfill = STATUS_FILL.get(SEV_LABEL.get(p["worst_severity"]))
        if wfill:
            ws.cell(row=i, column=9).fill = PatternFill("solid", fgColor=wfill)
    style_header_block(ws, 2, len(DASH_HEADERS))
    autofit(ws, [22, 9, 7, 7, 8, 9, 10, 20, 18])

    # Per-page tabs
    for p in pages:
        ws = wb.create_sheet(title=excel_title(p["page"], used))
        ws["A1"] = f"{p['page']}  ·  health {p['health_pct']}%"
        for c, h in enumerate(PAGE_HEADERS, start=1):
            ws.cell(row=2, column=c, value=h)
        status_col = PAGE_HEADERS.index("Status") + 1
        row = 3
        for r in p["rows"]:
            ws.append([
                r["name"], r["scope"], str(r["value"]), r["status"],
                r["fresh_field"], fmt_ts(r["freshness_ts"]),
                fmt_ts(r["last_changed"]), r["detail"],
            ])
            fill = STATUS_FILL.get(r["status"])
            if fill:
                ws.cell(row=row, column=status_col).fill = PatternFill("solid", fgColor=fill)
            row += 1
        style_header_block(ws, 2, len(PAGE_HEADERS))
        autofit(ws, [30, 16, 26, 18, 26, 19, 19, 50])

    return wb


# ---- drive upload -------------------------------------------------------------
def find_existing(drive):
    # Known sheet first — folder queries can miss under the service account,
    # which only sees files explicitly shared with it.
    try:
        f = drive.files().get(fileId=KNOWN_SHEET_ID,
                              fields="id,trashed,capabilities(canEdit)").execute()
        if not f.get("trashed") and f.get("capabilities", {}).get("canEdit"):
            return KNOWN_SHEET_ID
    except Exception:
        pass
    q = (f"name='{SHEET_NAME}' and '{FOLDER_ID}' in parents "
         f"and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false")
    hits = drive.files().list(q=q, fields="files(id,name)", supportsAllDrives=True,
                              includeItemsFromAllDrives=True).execute().get("files", [])
    return hits[0]["id"] if hits else None


def upload(drive):
    media = MediaFileUpload(XLSX_PATH, mimetype=XLSX_MIME, resumable=False)
    ssid = find_existing(drive)
    if ssid:
        # Also carries the 2026-07-22 rename (was "Main Site Health") onto the
        # existing file — cheap metadata update, same call as the content sync.
        drive.files().update(fileId=ssid, body={"name": SHEET_NAME}, media_body=media,
                             supportsAllDrives=True).execute()
    else:
        body = {"name": SHEET_NAME, "parents": [FOLDER_ID],
                "mimeType": "application/vnd.google-apps.spreadsheet"}
        ssid = drive.files().create(body=body, media_body=media, fields="id",
                                    supportsAllDrives=True).execute()["id"]
    return ssid


# ---- alert --------------------------------------------------------------------
def maybe_alert(pages, totals, url, now_utc, minisite_results=None):
    """Mini-site had zero Telegram alerting before this merge (2026-07-22) —
    its breaches are folded into the same digest here rather than a second,
    separate message, so one breach-worthy night produces one alert."""
    breaches = totals.get("ERROR", 0) + totals.get("STALE", 0) + totals.get("MISSING", 0)
    mini_breaches = []
    for r in (minisite_results or []):
        c = r["counts"]
        b = c.get("ERROR", 0) + c.get("MISSING", 0) + c.get("STALE", 0)
        if b:
            mini_breaches.append((r, b))
    if breaches == 0 and not mini_breaches:
        return
    try:
        from telegram_notify import send_message, TelegramSendError
    except Exception as e:
        print(f"(telegram unavailable: {e})")
        return
    total_breaches = breaches + sum(b for _, b in mini_breaches)
    lines = [f"⚠️ Fields Systems Health — {total_breaches} data point(s) need attention "
             f"({now_utc.astimezone(hc.AEST):%Y-%m-%d %H:%M AEST})"]
    for p in pages:
        c = p["counts"]
        b = c.get("ERROR", 0) + c.get("STALE", 0) + c.get("MISSING", 0)
        if b:
            lines.append(f"• {p['page']}: {p['health_pct']}% "
                         f"(err {c.get('ERROR',0)} / stale {c.get('STALE',0)} / miss {c.get('MISSING',0)})")
    if mini_breaches:
        lines.append(f"• Mini-Site: {len(mini_breaches)} home(s) with breaches")
        for r, b in sorted(mini_breaches, key=lambda x: -x[1])[:5]:
            lines.append(f"   – {r['address']} ({r['suburb']}): {r['health_pct']}%, {b} breach(es)")
        if len(mini_breaches) > 5:
            lines.append(f"   – …and {len(mini_breaches) - 5} more (see Mini-Site Dashboard tab)")
    lines.append(url)
    try:
        send_message("\n".join(lines), parse_mode="")
    except TelegramSendError as e:
        print(f"(telegram send failed: {e} — check TELEGRAM_BOT_TOKEN/CHAT_ID)")


# ---- main ---------------------------------------------------------------------
def set_env_from_file():
    if os.environ.get("COSMOS_CONNECTION_STRING"):
        return
    env_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), ".env")
    if os.path.exists(env_path):
        for line in open(env_path):
            if "=" in line and not line.startswith("#"):
                key, _, val = line.partition("=")
                os.environ.setdefault(key.strip(), val.strip().strip('"'))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--no-snapshot", action="store_true")
    ap.add_argument("--no-alert", action="store_true")
    ap.add_argument("--no-minisite", action="store_true",
                    help="skip mini-site tabs (debugging only)")
    args = ap.parse_args()

    set_env_from_file()
    pages, now_utc, totals = hc.run_audit(persist=not args.no_snapshot)
    wb = build_workbook(pages, now_utc, totals)

    minisite_results = []
    if not args.no_minisite:
        try:
            import minisite_health_check as mhc
            import minisite_health_to_sheet as mhs
            minisite_results, _ = mhc.run_audit(persist=not args.no_snapshot)
            mhs.build_workbook(minisite_results, now_utc, wb=wb)
        except Exception as e:
            print(f"(mini-site tabs skipped: {e})")

    wb.save(XLSX_PATH)
    drive = get_drive()
    ssid = upload(drive)
    url = f"https://docs.google.com/spreadsheets/d/{ssid}/edit"
    print(f"Synced {len(pages)} main-site pages + {len(minisite_results)} mini-site homes → {url}")
    if not args.no_alert:
        maybe_alert(pages, totals, url, now_utc, minisite_results=minisite_results)


def _alert_own_crash(exc):
    """Last-resort direct alert if main() itself dies before reaching
    maybe_alert() — e.g. a Mongo timeout, an unhandled bug in a new collector,
    a Drive API auth failure. Without this, a crash here means the sheet
    silently stops updating and NOTHING says so — the exact failure class
    this whole system was built to catch, one level up in the checker itself
    (flagged 2026-07-22: the only thing checking this job before was a
    self-referential Process Registry row that just said "this sheet,
    self-evidently running" — a no-op that assumes what it should verify).
    Deliberately does not import anything from hc/mhc/build_workbook — those
    are exactly what might have just crashed."""
    try:
        from telegram_notify import send_message, TelegramSendError
        try:
            send_message(
                f"🔴 Fields Systems Health FAILED to complete — sheet may not have updated.\n"
                f"{type(exc).__name__}: {exc}\n"
                f"Check logs/main-site-health.log on the VM.",
                parse_mode="")
        except TelegramSendError as e:
            print(f"(crash alert also failed to send: {e})")
    except Exception as e:
        print(f"(could not even attempt a crash alert: {e})")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        _alert_own_crash(e)
        raise
